import re
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import os
import openpyxl
from playwright.sync_api import sync_playwright
import time
import random
import string

app = Flask(__name__, static_folder='build', static_url_path='')
CORS(app)

# Platform patterns
PLATFORM_PATTERNS = {
    "DoubleClick": r"doubleclick\.net|securepubads\.g\.doubleclick\.net|tpc\.googlesyndication\.com",
    "Google Ads": r"googleadservices\.com|adservice\.google\.com|pagead2\.googlesyndication\.com",
    "Facebook": r"connect\.facebook\.net|graph\.facebook\.com|www\.facebook\.com",
    "LinkedIn": r"snap\.licdn\.com|www\.linkedin\.com|dc\.ads\.linkedin\.com",
    "TikTok": r"analytics\.tiktok\.com|business\.tiktok\.com|ads\.tiktok\.com",
    "Bing": r"bat\.bing\.com|c\.bing\.com",
    "Microsoft": r"analytics\.ms|dc\.services\.visualstudio\.com|c\.microsoft\.com",
    "Clarity": r"c\.clarity\.ms|clarity\.ms",
    "Hotjar": r"script\.hotjar\.com|insights\.hotjar\.com",
    "Full Story": r"fullstory\.com|edge\.fullstory\.com",
    "GA4": r"google-analytics\.com|analytics\.google\.com|ga4\.googleapis\.com",
    "Amazon": r"fls-na\.amazon\.com|amazon-adsystem\.com|aax\.amazon-adsystem\.com"
}

# Data patterns
DATA_PATTERNS = {
    "Page": r"(?:[?&]page=)([^&]+)",
    "User Query": r"(?:[?&](q|query|search)=)([^&]+)",
    "Referring Page (dr/ref)": r"(?:[?&](?:dr|ref)=)([^&]+)",  # Referring page (matches 'dr=' or 'ref=')
    "Current Page (dl/url)": r"(?:[?&](?:dl|url)=)([^&]+)",     # Current page (matches 'dl=' or 'url=')
    "Device ID (did)": r"(?:[?&]did=)([^&]+)"
}

# Function to generate a random string
def generate_random_string(length=8):
    return ''.join(random.choices(string.ascii_lowercase, k=length))

# Function to create or open an Excel file for logging
def create_excel_file(filename):
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = 'initial_requests'
        ws1.append(['Type', 'Method', 'URL', 'Platform', 'Data Shared'])  # Added new columns

        ws2 = wb.create_sheet(title='cookies')
        ws2.append(['Name', 'Value', 'Domain', 'Path', 'Expires', 'HttpOnly', 'Secure', 'SameSite'])

        wb.save(filename)

# Function to log data to Excel file
def log_to_excel(filename, sheet_name, row_data):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheet_name]
    sheet.append(row_data)
    wb.save(filename)

# Function to identify the platform based on the URL
def identify_platform(url):
    for platform, pattern in PLATFORM_PATTERNS.items():
        if re.search(pattern, url, re.IGNORECASE):
            return platform
    return None

# Function to identify data shared in the URL or request body
def identify_data_shared(url, request_body):
    shared_data = []
    for data_type, pattern in DATA_PATTERNS.items():
        if re.search(pattern, url, re.IGNORECASE) or (request_body and re.search(pattern, request_body, re.IGNORECASE)):
            shared_data.append(data_type)
    return ', '.join(shared_data) if shared_data else None

# Function to log cookies to Excel
def log_cookies_to_excel(page, filename):
    cookies = page.context.cookies()
    for cookie in cookies:
        cookie_details = [
            cookie.get('name'),
            cookie.get('value'),
            cookie.get('domain'),
            cookie.get('path'),
            cookie.get('expires'),
            cookie.get('httpOnly'),
            cookie.get('secure'),
            cookie.get('sameSite')
        ]
        log_to_excel(filename, 'cookies', cookie_details)

# Scraping function
def scrape_site(url):
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        # Increase timeout to 60 seconds and use 'networkidle' for dynamic content
        page.goto(url, timeout=60000)
        page.wait_for_load_state('networkidle')  # Wait until the page's network activity is idle

        # Define the updated search button selector
        search_button_selector = 'button[aria-label="Search"], button.btn-primary'

        try:
            # Attempt to find the button and ensure it's visible
            search_button = page.wait_for_selector(search_button_selector, timeout=30000)
            if search_button:
                search_button.click()
                print("Clicked on the 'Search' button or link.")

                # Wait for network idle after clicking the button to ensure dynamic content is loaded
                # page.wait_for_load_state('networkidle')  # Wait again for any dynamic content after clicking

                # Now create the Excel file and start logging
                create_excel_file('requests_and_cookies.xlsx')
                log_cookies_to_excel(page, 'requests_and_cookies.xlsx')

                # Handle requests as they are made
                def handle_request(request):
                    print(f"Captured: {request.method} {request.url}")  # Debugging step
                
                # page.on("request", handle_request)
                # page.on("requestfinished", lambda request: print(f"Request finished: {request.method} {request.url}"))
                # page.on("requestfailed", lambda request: print(f"Request failed: {request.method} {request.url}"))

                    platform = identify_platform(request.url)

                    # For GET requests, use the query parameters from the URL
                    if request.method == 'GET':
                        data_shared = identify_data_shared(request.url, None)  # For GET requests, use URL data (query parameters)
                    else:
                        # For POST, PUT, PATCH, DELETE, etc., use the request body (post_data)
                        data_shared = identify_data_shared(request.url, request.post_data)

                    # Capture all requests, regardless of the method (GET, POST, PUT, DELETE, etc.)
                    log_to_excel('requests_and_cookies.xlsx', 'initial_requests', [
                        'Request', request.method, request.url, platform, data_shared
                    ])

                page.on('request', handle_request)

                # Wait and check periodically for requests to complete
                page.wait_for_timeout(60000)  # Wait for 60 seconds or adjust as necessary to ensure we capture all requests

            else:
                print("No 'Search' button found.")

        except TimeoutError:
            print(f"Timeout: Unable to find the 'Search' button within 60 seconds.")
        
        page.close()
        browser.close()

@app.route('/scrape', methods=['GET', 'POST', 'PUT', 'DELETE'])
def scrape():
    # Initialize url variable to None
    url = None

    if request.method == 'POST':
        # For POST, get the URL from the JSON body
        data = request.json
        url = data.get('url')
    elif request.method == 'GET':
        # For GET, get the URL from query parameters
        url = request.args.get('url')
    elif request.method == 'PUT':
        # For PUT, get the URL from the JSON body
        data = request.json
        url = data.get('url')
    elif request.method == 'DELETE':
        # For DELETE, get the URL from query parameters
        url = request.args.get('url')

    # If no URL is provided, return an error
    if not url:
        return jsonify({"error": "URL is required"}), 400

    # Proceed with scraping the URL
    scrape_site(url)

    return jsonify({"message": "Scraping completed successfully!"})

# Serve React build files in production
@app.route('/')
def serve_react_app():
    return send_from_directory('build', 'index.html')

@app.route('/<path:path>')
def serve_react_static(path):
    return send_from_directory('build', path)

if __name__ == "__main__":
    app.run(debug=True, host='0.0.0.0', port=5001)
